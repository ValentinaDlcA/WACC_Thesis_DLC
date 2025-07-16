import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import os


# ------------------------------------------------------------
# 1. Load the projection table with Region & IncomeLevel
# ------------------------------------------------------------
proj_path = (
    "/Users/valentinadlc/Library/Caches/JetBrains/PyCharm2025.1/"
    "demo/PyCharmLearningProject/WACC Thesis/wacc_projection_with_regions_income.csv"
)
df = pd.read_csv(proj_path)

# Clean column names just in case
df.columns = [c.strip() for c in df.columns]
df["Year"] = df["Year"].astype(int)

# Create output folder for plots
output_folder = "/Users/valentinadlc/Library/Caches/JetBrains/PyCharm2025.1/demo/PyCharmLearningProject/WACC Thesis/plots"
os.makedirs(output_folder, exist_ok=True)

# ------------------------------------------------------------
# 2. LOOP over Technologies — for plots only!
# ------------------------------------------------------------
technologies = df["Technology"].unique()

for technology in technologies:
    print(f"\n=== Processing Technology: {technology} ===")

    df_tech = df[df["Technology"] == technology].copy()

    # ---- PLOT: For each Scenario — one plot per Scenario, all Regions ----
    scenarios = df_tech["Scenario"].unique()

    for scen in scenarios:
        df_scen = df_tech[df_tech["Scenario"] == scen].copy()

        regional_wacc = (
            df_scen.groupby(["Year", "Region"])["wacc_projection"]
            .mean()
            .reset_index()
            .pivot(index="Year", columns="Region", values="wacc_projection")
        )

        fig, ax = plt.subplots(figsize=(12, 7))
        regional_wacc.plot(ax=ax, linewidth=2.5, marker='o')

        ax.set_ylim(0.05, None)
        ax.yaxis.set_major_locator(mtick.MultipleLocator(0.01))
        ax.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.2f'))

        ax.set_xlabel("Year")
        ax.set_ylabel("Projected WACC")
        ax.set_title(f"{technology} – WACC projection  •  Scenario: {scen}")
        ax.grid(axis='y', linestyle='--', alpha=0.4)

        ax.legend(title="Region", bbox_to_anchor=(1.02, 1), loc="upper left", fontsize="medium")

        plt.tight_layout()

        # Save plot to PNG
        filename = f"{output_folder}/{technology}_Scenario_{scen}_by_Region.png"
        plt.savefig(filename)
        print(f"✅ Saved: {filename}")

        plt.close()

    # ---- PLOT: For each Region — one plot per Region, all Scenarios ----
    regions = df_tech["Region"].unique()

    for region in regions:
        df_region = df_tech[df_tech["Region"] == region].copy()

        scenario_wacc = (
            df_region.groupby(["Year", "Scenario"])["wacc_projection"]
            .mean()
            .reset_index()
            .pivot(index="Year", columns="Scenario", values="wacc_projection")
        )

        fig, ax = plt.subplots(figsize=(12, 7))
        scenario_wacc.plot(ax=ax, linewidth=2.5, marker='o')

        ax.set_ylim(0.05, None)
        ax.yaxis.set_major_locator(mtick.MultipleLocator(0.01))
        ax.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.2f'))

        ax.set_xlabel("Year")
        ax.set_ylabel("Projected WACC")
        ax.set_title(f"{technology} – WACC projection  •  Region: {region}")
        ax.grid(axis='y', linestyle='--', alpha=0.4)

        ax.legend(title="Scenario", bbox_to_anchor=(1.02, 1), loc="upper left", fontsize="medium")

        plt.tight_layout()

        # Save plot to PNG
        filename = f"{output_folder}/{technology}_Region_{region}_by_Scenario.png"
        plt.savefig(filename)
        print(f"✅ Saved: {filename}")

        plt.close()

    # ---- PLOT: For each Scenario — plot by IncomeLevel ----
    for scen in scenarios:
        df_scen = df_tech[df_tech["Scenario"] == scen].copy()

        income_wacc = (
            df_scen.groupby(["Year", "IncomeLevel"])["wacc_projection"]
            .mean()
            .reset_index()
            .pivot(index="Year", columns="IncomeLevel", values="wacc_projection")
        )

        fig, ax = plt.subplots(figsize=(12, 7))
        income_wacc.plot(ax=ax, linewidth=2.5, marker='o')

        ax.set_ylim(0.05, None)
        ax.yaxis.set_major_locator(mtick.MultipleLocator(0.01))
        ax.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.2f'))

        ax.set_xlabel("Year")
        ax.set_ylabel("Projected WACC")
        ax.set_title(f"{technology} – WACC projection by IncomeLevel  •  Scenario: {scen}")
        ax.grid(axis='y', linestyle='--', alpha=0.4)

        ax.legend(title="IncomeLevel", bbox_to_anchor=(1.02, 1), loc="upper left", fontsize="medium")

        plt.tight_layout()

        # Save plot to PNG
        filename = f"{output_folder}/{technology}_Scenario_{scen}_by_IncomeLevel.png"
        plt.savefig(filename)
        print(f"✅ Saved: {filename}")

        plt.close()

# ------------------------------------------------------------
# 3. SUMMARY TABLE — all Technologies!
# ------------------------------------------------------------

# ⚠️ IMPORTANT: Now use **full df**, not df_tech!
key_years = [2025, 2030, 2050]

summary_table = (
    df.loc[df["Year"].isin(key_years)]
    .groupby(["Technology", "Scenario", "Region", "Year"])["wacc_projection"]
    .mean()
    .round(4)
    .unstack("Year")
    .sort_index()
)

print("\nSummary table — Avg WACC by Technology, Region & Scenario:")
print(summary_table)

# Save to CSV
summary_table.to_csv(f"{output_folder}/wacc_summary_table_by_technology.csv")
summary_table.to_csv(
    r"/Users/valentinadlc/Library/Caches/JetBrains/PyCharm2025.1/demo/PyCharmLearningProject/WACC Thesis/Excel tables/wacc_summary_table_by_technology.csv")
print(f"✅ Saved: wacc_summary_table_by_technology.csv")


# ------------------------------------------------------------
# 4. SUMMARY TABLE — by IncomeLevel
# ------------------------------------------------------------
summary_income = (
    df.loc[df["Year"].isin(key_years)]
    .groupby(["Technology", "Scenario", "IncomeLevel", "Year"])["wacc_projection"]
    .mean()
    .round(4)
    .unstack("Year")
    .sort_index()
)

print("\nSummary table — Avg WACC by Technology, IncomeLevel & Scenario:")
print(summary_income)

# Make sure folder exists:
os.makedirs(r"/Users/valentinadlc/Library/Caches/JetBrains/PyCharmLearningProject/WACC Thesis/Excel tables", exist_ok=True)

# Save to CSV
summary_income.to_csv(f"{output_folder}/wacc_summary_table_by_incomelevel.csv")
summary_income.to_csv(
    r"/Users/valentinadlc/Library/Caches/JetBrains/PyCharmLearningProject/WACC Thesis/Excel tables/wacc_summary_table_by_incomelevel.csv")
print(f"✅ Saved: wacc_summary_table_by_incomelevel.csv")

#TO EXCEL
# ---- Folder path ----
excel_folder = r"/Users/valentinadlc/Library/Caches/JetBrains/PyCharm2025.1/demo/PyCharmLearningProject/WACC Thesis/Excel tables"
os.makedirs(excel_folder, exist_ok=True)

# ---- Load CSV ----
csv_path = f"{excel_folder}/wacc_summary_table_by_technology.csv"
summary_table = pd.read_csv(csv_path)

# ---- Save to Excel ----
excel_path = f"{excel_folder}/wacc_summary_table_by_technology.xlsx"
summary_table.to_excel(excel_path, index=False)

# ---- Apply formatting ----
wb = load_workbook(excel_path)
ws = wb.active

# Bold header + center alignment + fill
header_font = Font(bold=True)
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

for col_num, cell in enumerate(ws[1], 1):
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = header_fill
    ws.column_dimensions[get_column_letter(col_num)].width = 22

# Center-align all data
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Save again
wb.save(excel_path)
print(f"✅ Saved formatted Excel: {excel_path}")